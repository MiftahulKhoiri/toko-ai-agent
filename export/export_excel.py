"""export/export_excel.py - Export laporan ke file Excel (.xlsx) untuk toko-ai-agent

Fitur:
- Export laporan harian
- Export laporan bulanan
- Export laporan stok
- File Excel siap dibuka
"""

from datetime import datetime
from pathlib import Path
from typing import List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font

from sqlalchemy import select, func, extract
from sqlalchemy.exc import SQLAlchemyError

from database.db import SessionLocal
from database.models import (
    Transaksi,
    Biaya,
    Barang,
    StokAudit,
)

from config import TODAY


# =========================================================
# CONFIG
# =========================================================

EXPORT_DIR = Path("export")

try:
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
except Exception as exc:
    print(f"[ERROR] Gagal membuat folder export: {exc}")


# =========================================================
# UTIL
# =========================================================

def get_session():
    return SessionLocal()


def get_timestamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def get_tanggal():
    return datetime.strptime(TODAY, "%Y-%m-%d").date()


def set_header(sheet, headers: List[str]):
    """Set header Excel dengan bold font"""
    bold_font = Font(bold=True)
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = bold_font


# =========================================================
# EXPORT HARIAN
# =========================================================

def export_laporan_harian() -> None:
    session = get_session()
    try:
        tanggal = get_tanggal()
        pendapatan = session.execute(
            select(func.sum(Transaksi.pendapatan))
            .where(Transaksi.tanggal == tanggal)
        ).scalar()
        biaya = session.execute(
            select(func.sum(Biaya.jumlah))
            .where(Biaya.tanggal == tanggal)
        ).scalar()
        total_pendapatan = float(pendapatan or 0)
        total_biaya = float(biaya or 0)
        laba = total_pendapatan - total_biaya

        wb = Workbook()
        ws = wb.active
        ws.title = "Laporan Harian"

        headers = ["Tanggal", "Pendapatan", "Biaya", "Laba"]
        set_header(ws, headers)
        ws.append([str(tanggal), total_pendapatan, total_biaya, laba])

        filename = EXPORT_DIR / f"laporan_harian_{get_timestamp()}.xlsx"
        wb.save(filename)
        print(f"Export berhasil: {filename}")
    except SQLAlchemyError as exc:
        print(f"[ERROR] Gagal export laporan harian: {exc}")
    except Exception as exc:
        print(f"[ERROR] Export error: {exc}")
    finally:
        session.close()


# =========================================================
# EXPORT BULANAN
# =========================================================

def export_laporan_bulanan() -> None:
    session = get_session()
    try:
        tanggal = get_tanggal()
        bulan = tanggal.month
        tahun = tanggal.year

        pendapatan = session.execute(
            select(func.sum(Transaksi.pendapatan))
            .where(
                extract("month", Transaksi.tanggal) == bulan,
                extract("year", Transaksi.tanggal) == tahun,
            )
        ).scalar()
        biaya = session.execute(
            select(func.sum(Biaya.jumlah))
            .where(
                extract("month", Biaya.tanggal) == bulan,
                extract("year", Biaya.tanggal) == tahun,
            )
        ).scalar()

        total_pendapatan = float(pendapatan or 0)
        total_biaya = float(biaya or 0)
        laba = total_pendapatan - total_biaya

        wb = Workbook()
        ws = wb.active
        ws.title = "Laporan Bulanan"

        headers = ["Bulan", "Tahun", "Pendapatan", "Biaya", "Laba"]
        set_header(ws, headers)
        ws.append([bulan, tahun, total_pendapatan, total_biaya, laba])

        filename = EXPORT_DIR / f"laporan_bulanan_{get_timestamp()}.xlsx"
        wb.save(filename)
        print(f"Export berhasil: {filename}")
    except SQLAlchemyError as exc:
        print(f"[ERROR] Gagal export laporan bulanan: {exc}")
    except Exception as exc:
        print(f"[ERROR] Export error: {exc}")
    finally:
        session.close()


# =========================================================
# EXPORT STOK
# =========================================================

def export_laporan_stok() -> None:
    session = get_session()
    try:
        tanggal = get_tanggal()
        results = session.execute(
            select(
                Barang.nama,
                StokAudit.stok_awal,
                StokAudit.stok_masuk,
                StokAudit.stok_keluar,
                StokAudit.stok_akhir,
            )
            .join(Barang, Barang.id == StokAudit.barang_id)
            .where(StokAudit.tanggal == tanggal)
        ).all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Laporan Stok"

        headers = ["Nama Barang", "Stok Awal", "Masuk", "Keluar", "Stok Akhir"]
        set_header(ws, headers)

        if not results:
            print("Tidak ada data stok")
        else:
            for row in results:
                ws.append([
                    row.nama,
                    row.stok_awal,
                    row.stok_masuk,
                    row.stok_keluar,
                    row.stok_akhir,
                ])

        filename = EXPORT_DIR / f"laporan_stok_{get_timestamp()}.xlsx"
        wb.save(filename)
        print(f"Export berhasil: {filename}")
    except SQLAlchemyError as exc:
        print(f"[ERROR] Gagal export laporan stok: {exc}")
    except Exception as exc:
        print(f"[ERROR] Export error: {exc}")
    finally:
        session.close()


# =========================================================
# TEST
# =========================================================

if __name__ == "__main__":
    print("TEST EXPORT EXCEL")
    export_laporan_harian()
    export_laporan_bulanan()
    export_laporan_stok()